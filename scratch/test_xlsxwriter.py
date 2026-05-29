import xlsxwriter
import openpyxl

workbook = xlsxwriter.Workbook('test.xlsx')
worksheet = workbook.add_worksheet()

# Test merging single cell
worksheet.merge_range('A1:A1', 'Test Single', None)
# Test merging two cells
worksheet.merge_range('B1:C1', 'Test Multi', None)

workbook.close()

# Load and check values
wb = openpyxl.load_workbook('test.xlsx')
ws = wb.active
print("A1 value:", ws['A1'].value)
print("B1 value:", ws['B1'].value)
print("C1 value:", ws['C1'].value)
